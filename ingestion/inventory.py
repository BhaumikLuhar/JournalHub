from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Project paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[1]

RAW_DIR = PROJECT_ROOT / "data" / "raw"
SCIMAGO_DIR = RAW_DIR / "scimago"
ABDC_DIR = RAW_DIR / "abdc"
ABS_DIR = RAW_DIR / "abs"
REPEC_DIR = RAW_DIR / "repec"
FT50_DIR = RAW_DIR / "ft50"

REPORTS_DIR = PROJECT_ROOT / "reports"
OUTPUT_FILE = REPORTS_DIR / "data_dictionary.csv"


# ---------------------------------------------------------------------------
# Inventory schema
# ---------------------------------------------------------------------------

INVENTORY_COLUMNS = [
    "source",
    "file_name",
    "sheet_name",
    "header_row_index",
    "delimiter",
    "encoding",
    "row_count",
    "column_list",
    "identifier_status",
    "notes",
]


# ---------------------------------------------------------------------------
# General helpers
# ---------------------------------------------------------------------------

def clean_cell(value: Any) -> str:
    """Convert a cell value into a clean string for inventory metadata."""
    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    return str(value).strip()


def columns_to_json(columns: list[str]) -> str:
    """Store the exact column list as JSON inside the CSV report."""
    return json.dumps(columns, ensure_ascii=False)


def add_record(
    records: list[dict[str, Any]],
    *,
    source: str,
    file_name: str,
    sheet_name: str | None,
    header_row_index: int | None,
    delimiter: str | None,
    encoding: str | None,
    row_count: int,
    columns: list[str],
    identifier_status: str,
    notes: str,
) -> None:
    """Append one normalized inventory record."""
    records.append(
        {
            "source": source,
            "file_name": file_name,
            "sheet_name": sheet_name or "",
            "header_row_index": (
                "" if header_row_index is None else header_row_index
            ),
            "delimiter": delimiter or "",
            "encoding": encoding or "",
            "row_count": row_count,
            "column_list": columns_to_json(columns),
            "identifier_status": identifier_status,
            "notes": notes,
        }
    )


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def read_csv_with_encoding_fallback(
    path: Path,
    delimiter: str | None = None,
) -> tuple[pd.DataFrame, str, str]:
    """
    Read a CSV while preserving the source structure.

    Returns:
        dataframe,
        encoding_used,
        delimiter_used
    """
    encodings = ["utf-8", "latin-1"]

    for encoding in encodings:
        try:
            if delimiter is not None:
                df = pd.read_csv(
                    path,
                    sep=delimiter,
                    encoding=encoding,
                )
                return df, encoding, delimiter

            df = pd.read_csv(
                path,
                encoding=encoding,
            )

            return df, encoding, ","

        except UnicodeDecodeError:
            continue

    raise UnicodeDecodeError(
        "csv",
        b"",
        0,
        1,
        f"Unable to decode CSV file: {path}",
    )


def detect_csv_delimiter(path: Path, encoding: str) -> str:
    """
    Detect the delimiter of a CSV file without transforming the file.

    The inventory task expects SCImago to use ';' and the other standard
    CSV files to use ','.
    """
    with path.open("r", encoding=encoding, newline="") as file:
        sample = file.read(8192)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
        return dialect.delimiter
    except csv.Error:
        return ","


def inventory_csv(
    records: list[dict[str, Any]],
    path: Path,
    source: str,
    *,
    expected_delimiter: str | None = None,
    identifier_status: str = "",
    notes: str = "",
) -> None:
    """Inventory one CSV file."""
    # First determine encoding.
    encoding = "utf-8"

    try:
        with path.open("r", encoding="utf-8", newline="") as file:
            file.read(4096)
    except UnicodeDecodeError:
        encoding = "latin-1"

    # Determine delimiter.
    if expected_delimiter is not None:
        delimiter = expected_delimiter
    else:
        delimiter = detect_csv_delimiter(path, encoding)

    # Read using the detected/expected delimiter.
    df, encoding_used, delimiter_used = read_csv_with_encoding_fallback(
        path,
        delimiter=delimiter,
    )

    # The sample is deliberately collected but not written into the final
    # schema because the Day 1 data_dictionary specification does not have
    # a sample column. We print it to the console for verification.
    sample_rows = df.head(2).to_dict(orient="records")

    print()
    print(f"CSV: {path}")
    print(f"  Encoding: {encoding_used}")
    print(f"  Delimiter: {repr(delimiter_used)}")
    print(f"  Rows: {len(df)}")
    print(f"  Columns: {list(df.columns)}")
    print(f"  First 2 rows: {sample_rows}")

    add_record(
        records,
        source=source,
        file_name=str(path.relative_to(RAW_DIR)),
        sheet_name=None,
        header_row_index=0,
        delimiter=delimiter_used,
        encoding=encoding_used,
        row_count=len(df),
        columns=[str(column) for column in df.columns],
        identifier_status=identifier_status,
        notes=notes,
    )


# ---------------------------------------------------------------------------
# ABDC helpers
# ---------------------------------------------------------------------------

ABDC_HEADER_LABELS = {
    "journal title",
    "journal name",
}


def find_abdc_header_row(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    max_scan_rows: int = 15,
) -> int | None:
    """
    Find the first row containing 'Journal Title' or 'Journal Name'.

    Returns a 0-indexed row number, matching the project plan.
    """
    for row_index, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=max_scan_rows,
            values_only=True,
        )
    ):
        for value in row:
            normalized = clean_cell(value).casefold()

            if normalized in ABDC_HEADER_LABELS:
                return row_index

    return None


def inventory_abdc(
    records: list[dict[str, Any]],
) -> None:
    """Inventory every sheet in the ABDC workbook."""
    workbooks = sorted(ABDC_DIR.glob("*.xlsx"))

    if not workbooks:
        raise FileNotFoundError(
            f"No ABDC .xlsx workbook found in {ABDC_DIR}"
        )

    if len(workbooks) > 1:
        raise RuntimeError(
            f"Expected one ABDC workbook, found {len(workbooks)}: "
            f"{[path.name for path in workbooks]}"
        )

    path = workbooks[0]

    print()
    print(f"ABDC workbook: {path}")

    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            header_row = find_abdc_header_row(worksheet)

            if header_row is None:
                add_record(
                    records,
                    source="ABDC",
                    file_name=path.name,
                    sheet_name=sheet_name,
                    header_row_index=None,
                    delimiter=None,
                    encoding=None,
                    row_count=0,
                    columns=[],
                    identifier_status="UNKNOWN",
                    notes=(
                        "No 'Journal Title' or 'Journal Name' header "
                        "found within the first 15 rows."
                    ),
                )
                continue

            # openpyxl's max_row can include formatting-only rows, so use
            # pandas to determine the actual data shape from the detected
            # header row.
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                header=header_row,
            )

            columns = [str(column) for column in df.columns]

            print()
            print(f"Sheet: {sheet_name}")
            print(f"  Header row (0-indexed): {header_row}")
            print(f"  Rows: {len(df)}")
            print(f"  Columns: {columns}")

            add_record(
                records,
                source="ABDC",
                file_name=path.name,
                sheet_name=sheet_name,
                header_row_index=header_row,
                delimiter=None,
                encoding=None,
                row_count=len(df),
                columns=columns,
                identifier_status="ISSN",
                notes=(
                    "Header row detected dynamically by scanning the "
                    "first 15 rows."
                ),
            )
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Source-specific inventory
# ---------------------------------------------------------------------------

def inventory_scimago(records: list[dict[str, Any]]) -> None:
    """
    Inventory all SCImago CSV files recursively.

    The actual JournalHub raw structure groups SCImago files by subject
    folder, so recursive discovery is intentional.
    """
    files = sorted(SCIMAGO_DIR.rglob("*.csv"))

    if not files:
        raise FileNotFoundError(
            f"No SCImago CSV files found recursively under {SCIMAGO_DIR}"
        )

    for path in files:
        inventory_csv(
            records,
            path,
            "SCIMAGO",
            expected_delimiter=";",
            identifier_status="ISSN + SCIMAGO_SOURCE_ID",
            notes=(
                "SCImago CSV; semicolon-delimited. "
                "Source folder is preserved as part of the raw-data layout."
            ),
        )


def inventory_abs(records: list[dict[str, Any]]) -> None:
    """Inventory the ABS/AJG CSV."""
    files = sorted(ABS_DIR.glob("*.csv"))

    if not files:
        raise FileNotFoundError(
            f"No ABS CSV found in {ABS_DIR}"
        )

    if len(files) > 1:
        raise RuntimeError(
            f"Expected one ABS CSV, found {len(files)}: "
            f"{[path.name for path in files]}"
        )

    inventory_csv(
        records,
        files[0],
        "ABS",
        expected_delimiter=",",
        identifier_status="ISSN",
        notes=(
            "ABS/AJG 2024 file. Expected wide format with rating "
            "columns AJG2024, AJG2021 and AJG2018."
        ),
    )


def inventory_repec(records: list[dict[str, Any]]) -> None:
    """Inventory the RePEc aggregate rankings CSV."""
    files = sorted(REPEC_DIR.glob("*.csv"))

    if not files:
        raise FileNotFoundError(
            f"No RePEc CSV found in {REPEC_DIR}"
        )

    if len(files) > 1:
        raise RuntimeError(
            f"Expected one RePEc CSV, found {len(files)}: "
            f"{[path.name for path in files]}"
        )

    inventory_csv(
        records,
        files[0],
        "REPEC",
        expected_delimiter=",",
        identifier_status="NONE",
        notes=(
            "No ISSN or stable ID expected. Title-only matching required; "
            "publisher is bundled into the journal display name."
        ),
    )


def inventory_ft50(records: list[dict[str, Any]]) -> None:
    """Inventory the FT50 CSV."""
    files = sorted(FT50_DIR.glob("*.csv"))

    if not files:
        raise FileNotFoundError(
            f"No FT50 CSV found in {FT50_DIR}"
        )

    if len(files) > 1:
        raise RuntimeError(
            f"Expected one FT50 CSV, found {len(files)}: "
            f"{[path.name for path in files]}"
        )

    inventory_csv(
        records,
        files[0],
        "FT50",
        expected_delimiter=",",
        identifier_status="NONE",
        notes=(
            "No ISSN. Expected columns are rank, journal_name, ft50_year. "
            "Title-only matching required."
        ),
    )


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_inventory(records: list[dict[str, Any]]) -> None:
    """Run basic sanity checks before writing the report."""
    if not records:
        raise RuntimeError("Inventory produced no records.")

    for record in records:
        if not record["file_name"]:
            raise RuntimeError("Inventory record has an empty file_name.")

        if record["row_count"] < 0:
            raise RuntimeError(
                f"Negative row count for {record['file_name']}"
            )

        if record["row_count"] > 0 and not record["column_list"]:
            raise RuntimeError(
                f"File has rows but no columns: {record['file_name']}"
            )

    print()
    print("Inventory sanity checks passed.")


# ---------------------------------------------------------------------------
# Report writer
# ---------------------------------------------------------------------------

def write_report(records: list[dict[str, Any]]) -> None:
    """Write the complete inventory report."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    dataframe = pd.DataFrame(
        records,
        columns=INVENTORY_COLUMNS,
    )

    dataframe.to_csv(
        OUTPUT_FILE,
        index=False,
        encoding="utf-8",
    )

    print()
    print(f"Inventory report written to: {OUTPUT_FILE}")
    print(f"Inventory records: {len(dataframe)}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 80)
    print("JournalHub — Day 1 Raw Data Inventory")
    print("=" * 80)

    print()
    print(f"Project root: {PROJECT_ROOT}")
    print(f"Raw data directory: {RAW_DIR}")

    if not RAW_DIR.exists():
        raise FileNotFoundError(
            f"Raw data directory does not exist: {RAW_DIR}"
        )

    records: list[dict[str, Any]] = []

    print()
    print("=== SCIMAGO ===")
    inventory_scimago(records)

    print()
    print("=== ABDC ===")
    inventory_abdc(records)

    print()
    print("=== ABS ===")
    inventory_abs(records)

    print()
    print("=== REPEC ===")
    inventory_repec(records)

    print()
    print("=== FT50 ===")
    inventory_ft50(records)

    validate_inventory(records)
    write_report(records)

    print()
    print("=" * 80)
    print("JournalHub inventory completed successfully.")
    print("=" * 80)


if __name__ == "__main__":
    main()